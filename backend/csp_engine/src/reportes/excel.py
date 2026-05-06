# =============================================================================
#  SmartCampusAI — src/reportes/excel.py
#  Exportación de resultados a Excel (.xlsx) con formato profesional
# =============================================================================

import json
import os
from collections import defaultdict, Counter

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from src.utils.horarios import parsear_horario

DIAS_SEMANA = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes"]

# ── Paleta de colores ─────────────────────────────────────────────────────────
C = {
    "azul_dark":  "1A3A5C",
    "azul_med":   "2E6DA4",
    "azul_light": "D6E4F0",
    "verde_dark": "1E6B3C",
    "verde_light":"D4EDDA",
    "rojo_dark":  "8B0000",
    "rojo_light": "FDECEA",
    "naranja_dk": "7B3F00",
    "naranja_lt": "FFF3CD",
    "gris":       "F5F5F5",
    "blanco":     "FFFFFF",
}
DIA_COLOR = {
    "Lunes":     "E3F2FD",
    "Martes":    "E8F5E9",
    "Miércoles": "FFF8E1",
    "Jueves":    "FCE4EC",
    "Viernes":   "F3E5F5",
}


def _borde():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def _hc(ws, row, col, val, bg="1A3A5C", fg="FFFFFF", sz=10, bold=True):
    """Celda de encabezado."""
    c = ws.cell(row=row, column=col, value=val)
    c.font      = Font(name="Arial", bold=bold, color=fg, size=sz)
    c.fill      = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border    = _borde()
    return c


def _dc(ws, row, col, val, bg="FFFFFF", fg="000000", bold=False, align="left", sz=9):
    """Celda de dato."""
    c = ws.cell(row=row, column=col, value=val)
    c.font      = Font(name="Arial", bold=bold, color=fg, size=sz)
    c.fill      = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal=align, vertical="center")
    c.border    = _borde()
    return c


def exportar_excel(resultado: dict, ruta_salida: str = "output/SmartCampusAI_Resultado.xlsx"):
    """
    Genera un archivo Excel con 5 hojas:
        1. 📊 Resumen        — KPIs y métricas del solver
        2. ✅ Clases Asignadas — tabla completa con filtros
        3. ❌ Sin Asignar     — clases no resueltas con motivo
        4. 🗓️ Horario Semanal — ocupación de salones por día
        5. 👨‍🏫 Horario Profesores — carga semanal por profesor
    """
    asignadas    = resultado["asignadas"]
    no_asignadas = resultado["no_asignadas"]
    total        = len(asignadas) + len(no_asignadas)
    pct_ok       = len(asignadas) / total * 100 if total > 0 else 0
    recup        = sum(1 for a in asignadas if a.get("recuperada_por_backtracking"))
    reasig       = sum(1 for a in asignadas if a.get("reasignada_por_backtracking"))
    desps        = [a["desperdicio"] for a in asignadas]
    avg_d        = sum(desps) / len(desps) if desps else 0

    wb = Workbook()

    # ── HOJA 1 — RESUMEN ─────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "📊 Resumen"
    ws1.sheet_view.showGridLines = False

    ws1.merge_cells("A1:I1")
    c = ws1["A1"]
    c.value     = "SmartCampusAI — Reporte de Asignación de Salones (Lunes–Viernes)"
    c.font      = Font(name="Arial", bold=True, size=15, color="FFFFFF")
    c.fill      = PatternFill("solid", fgColor=C["azul_dark"])
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 36

    ws1.merge_cells("A2:I2")
    c = ws1["A2"]
    c.value     = "Sistema CSP con Heurística MRV + Backtracking | Jornada 6:00–21:00"
    c.font      = Font(name="Arial", italic=True, size=10, color="FFFFFF")
    c.fill      = PatternFill("solid", fgColor=C["azul_med"])
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[2].height = 20
    ws1.row_dimensions[3].height = 10

    kpis = [
        ("TOTAL CLASES",  total,             "000000",        "EBF5FB"),
        ("✅ ASIGNADAS",   len(asignadas),    C["verde_dark"], C["verde_light"]),
        ("❌ SIN ASIGNAR", len(no_asignadas), C["rojo_dark"],  C["rojo_light"]),
        ("% ÉXITO",        f"{pct_ok:.1f}%", C["verde_dark"], C["verde_light"]),
        ("BACKTRACKING",   recup,             C["naranja_dk"], C["naranja_lt"]),
        ("DESPERDICIO Ø",  f"{avg_d:.1f}",   "000000",        C["gris"]),
    ]
    kpi_cols = [1, 2, 3, 4, 6, 7]
    for i, (label, val, fg, bg) in enumerate(kpis):
        col = kpi_cols[i]
        for r in [4, 5, 6]:
            ws1.cell(row=r, column=col).fill   = PatternFill("solid", fgColor=bg)
            ws1.cell(row=r, column=col).border = _borde()
        ws1.cell(row=4, column=col).value     = label
        ws1.cell(row=4, column=col).font      = Font(name="Arial", bold=True, size=8, color=fg)
        ws1.cell(row=4, column=col).alignment = Alignment(horizontal="center", vertical="center")
        ws1.cell(row=5, column=col).value     = val
        ws1.cell(row=5, column=col).font      = Font(name="Arial", bold=True, size=18, color=fg)
        ws1.cell(row=5, column=col).alignment = Alignment(horizontal="center", vertical="center")
    for r, h in [(4, 18), (5, 34), (6, 8)]:
        ws1.row_dimensions[r].height = h
    ws1.row_dimensions[7].height = 12

    metricas = [
        ("Fase 1 — Greedy MRV",         len(asignadas) - recup),
        ("Fase 2 — Backtracking",        recup),
        ("Clases reasignadas (BT)",       reasig),
        ("Desperdicio mínimo (sillas)",   min(desps) if desps else 0),
        ("Desperdicio máximo (sillas)",   max(desps) if desps else 0),
        ("Desperdicio promedio (sillas)", f"{avg_d:.1f}"),
    ]
    _hc(ws1, 8, 1, "Métrica",   bg=C["azul_med"])
    _hc(ws1, 8, 2, "Valor",     bg=C["azul_med"])
    _hc(ws1, 8, 4, "Día",       bg=C["azul_med"])
    _hc(ws1, 8, 5, "Asignadas", bg=C["azul_med"])
    ws1.row_dimensions[8].height = 18

    por_dia: dict[str, int] = defaultdict(int)
    for a in asignadas:
        por_dia[a["dia_asignado"]] += 1

    for ri, (label, val) in enumerate(metricas, 9):
        bg = C["gris"] if ri % 2 == 0 else C["blanco"]
        _dc(ws1, ri, 1, label, bg=bg, bold=True, sz=9)
        _dc(ws1, ri, 2, val,   bg=bg, align="center", sz=9)
        ws1.row_dimensions[ri].height = 16

    for ri, dia in enumerate(DIAS_SEMANA, 9):
        bg = DIA_COLOR.get(dia, C["blanco"])
        _dc(ws1, ri, 4, dia,          bg=bg, bold=True,   sz=9)
        _dc(ws1, ri, 5, por_dia[dia], bg=bg, align="center", sz=9)
        ws1.row_dimensions[ri].height = 16

    for col, w in [(1,28),(2,10),(3,4),(4,14),(5,12),(6,12),(7,12),(8,12),(9,12)]:
        ws1.column_dimensions[get_column_letter(col)].width = w

    # ── HOJA 2 — CLASES ASIGNADAS ────────────────────────────────────────────
    ws2 = wb.create_sheet("✅ Clases Asignadas")
    ws2.sheet_view.showGridLines = False

    ws2.merge_cells("A1:N1")
    c = ws2["A1"]
    c.value     = f"Clases Asignadas — {len(asignadas)} de {total} ({pct_ok:.1f}%)"
    c.font      = Font(name="Arial", bold=True, size=13, color="FFFFFF")
    c.fill      = PatternFill("solid", fgColor=C["verde_dark"])
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 26

    hdrs2 = ["#","Materia","Grupo","Profesor","Tipo","Día","Horario","Estudiantes",
             "Salón","Bloque","Cap.","Desperdicio","BT ↩","BT →"]
    for ci, h in enumerate(hdrs2, 1):
        _hc(ws2, 2, ci, h, bg=C["azul_med"])
    ws2.row_dimensions[2].height = 18

    sorted_a = sorted(asignadas, key=lambda x: (
        DIAS_SEMANA.index(x["dia_asignado"]),
        parsear_horario(x["horario"])[0],
        x["materia"]
    ))
    for ri, a in enumerate(sorted_a, 3):
        is_rec = a.get("recuperada_por_backtracking", False)
        is_rea = a.get("reasignada_por_backtracking", False)
        dia    = a["dia_asignado"]
        bg = DIA_COLOR.get(dia, C["blanco"])
        if is_rec: bg = "EAF7EA"
        if is_rea: bg = "FFF9E6"
        vals = [ri-2, a["materia"], a["grupo"], a["profesor"], a["tipo"],
                dia, a["horario"], a["estudiantes"],
                a["salon_asignado"], a["bloque_salon"], a["capacidad_salon"],
                a["desperdicio"],
                "✔" if is_rec else "", "✔" if is_rea else ""]
        for ci, v in enumerate(vals, 1):
            al = "center" if ci in [1,3,5,6,7,8,9,11,12,13,14] else "left"
            _dc(ws2, ri, ci, v, bg=bg, align=al, sz=9)
        ws2.row_dimensions[ri].height = 14

    for ci, w in enumerate([5,32,10,22,12,12,14,11,10,14,7,12,7,7], 1):
        ws2.column_dimensions[get_column_letter(ci)].width = w
    ws2.freeze_panes = "A3"
    ws2.auto_filter.ref = f"A2:N{2+len(sorted_a)}"

    # ── HOJA 3 — SIN ASIGNAR ─────────────────────────────────────────────────
    ws3 = wb.create_sheet("❌ Sin Asignar")
    ws3.sheet_view.showGridLines = False

    ws3.merge_cells("A1:K1")
    c = ws3["A1"]
    c.value     = f"Clases Sin Asignar — {len(no_asignadas)} (limitación real de equipamiento)"
    c.font      = Font(name="Arial", bold=True, size=13, color="FFFFFF")
    c.fill      = PatternFill("solid", fgColor=C["rojo_dark"])
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws3.row_dimensions[1].height = 26

    hdrs3 = ["#","Materia","Grupo","Profesor","Tipo","Horario","Estudiantes",
             "Req. VB","Req. PC","Req. Lab","Motivo"]
    for ci, h in enumerate(hdrs3, 1):
        _hc(ws3, 2, ci, h, bg=C["rojo_dark"])
    ws3.row_dimensions[2].height = 18

    for ri, n in enumerate(sorted(no_asignadas, key=lambda x: x["horario"]), 3):
        bg = "FFF5F5" if ri % 2 != 0 else "FDECEA"
        motivo = ("Salones con PC insuficientes — solo 6 disponibles en el campus"
                  if n.get("requiere_computadores")
                  else "Todos los salones válidos están ocupados en ese horario")
        vals = [ri-2, n["materia"], n["grupo"], n["profesor"], n["tipo"],
                n["horario"], n["estudiantes"],
                "Sí" if n.get("requiere_videobeam")    else "No",
                "Sí" if n.get("requiere_computadores") else "No",
                "Sí" if n.get("requiere_laboratorio")  else "No",
                motivo]
        for ci, v in enumerate(vals, 1):
            al = "center" if ci in [1,3,5,6,7,8,9,10] else "left"
            fg_c = C["rojo_dark"] if (ci == 9 and v == "Sí") else "000000"
            _dc(ws3, ri, ci, v, bg=bg, align=al, fg=fg_c, bold=(ci==9 and v=="Sí"), sz=9)
        ws3.row_dimensions[ri].height = 14

    for ci, w in enumerate([5,32,10,22,12,14,11,9,9,9,42], 1):
        ws3.column_dimensions[get_column_letter(ci)].width = w
    ws3.freeze_panes = "A3"
    ws3.auto_filter.ref = f"A2:K{2+len(no_asignadas)}"

    # ── HOJA 4 — HORARIO SEMANAL POR SALÓN ───────────────────────────────────
    ws4 = wb.create_sheet("🗓️ Horario Semanal")
    ws4.sheet_view.showGridLines = False

    ws4.merge_cells("A1:H1")
    c = ws4["A1"]
    c.value     = "Vista Semanal — Ocupación de Salones por Día y Horario"
    c.font      = Font(name="Arial", bold=True, size=13, color="FFFFFF")
    c.fill      = PatternFill("solid", fgColor=C["azul_dark"])
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws4.row_dimensions[1].height = 26

    for ci, h in enumerate(["Salón","Bloque","Cap.","Día","Horario","Materia","Grupo","Estudiantes"], 1):
        _hc(ws4, 2, ci, h, bg=C["azul_med"])
    ws4.row_dimensions[2].height = 18

    salon_dia: dict = defaultdict(list)
    for a in asignadas:
        salon_dia[a["salon_asignado"]].append(a)

    cur_row = 3
    for sid in sorted(salon_dia.keys()):
        cls     = salon_dia[sid]
        cap     = cls[0]["capacidad_salon"]
        bloque  = cls[0]["bloque_salon"]
        cls_ord = sorted(cls, key=lambda x: (
            DIAS_SEMANA.index(x["dia_asignado"]),
            parsear_horario(x["horario"])[0]
        ))
        for ci in range(1, 9):
            ws4.cell(row=cur_row, column=ci).fill   = PatternFill("solid", fgColor=C["azul_light"])
            ws4.cell(row=cur_row, column=ci).border = _borde()
        ws4.cell(row=cur_row, column=1).value     = sid
        ws4.cell(row=cur_row, column=1).font      = Font(name="Arial", bold=True, size=9, color=C["azul_dark"])
        ws4.cell(row=cur_row, column=1).alignment = Alignment(horizontal="center", vertical="center")
        ws4.cell(row=cur_row, column=2).value     = bloque
        ws4.cell(row=cur_row, column=2).font      = Font(name="Arial", bold=True, size=9, color=C["azul_dark"])
        ws4.cell(row=cur_row, column=2).alignment = Alignment(horizontal="center", vertical="center")
        ws4.cell(row=cur_row, column=3).value     = cap
        ws4.cell(row=cur_row, column=3).font      = Font(name="Arial", bold=True, size=9, color=C["azul_dark"])
        ws4.cell(row=cur_row, column=3).alignment = Alignment(horizontal="center", vertical="center")
        ws4.cell(row=cur_row, column=4).value     = f"{len(cls)} clase(s) esta semana"
        ws4.cell(row=cur_row, column=4).font      = Font(name="Arial", italic=True, size=9, color=C["azul_dark"])
        ws4.row_dimensions[cur_row].height = 15
        cur_row += 1
        for cl in cls_ord:
            dia = cl["dia_asignado"]
            bg  = DIA_COLOR.get(dia, C["blanco"])
            for ci, v in enumerate(["","","",dia,cl["horario"],cl["materia"],cl["grupo"],cl["estudiantes"]], 1):
                al = "center" if ci in [1,2,3,4,5,8] else "left"
                _dc(ws4, cur_row, ci, v, bg=bg, align=al, sz=9)
            ws4.row_dimensions[cur_row].height = 14
            cur_row += 1

    for ci, w in enumerate([10,16,7,12,14,32,10,12], 1):
        ws4.column_dimensions[get_column_letter(ci)].width = w
    ws4.freeze_panes = "A3"

    # ── HOJA 5 — HORARIO POR PROFESOR ────────────────────────────────────────
    ws5 = wb.create_sheet("👨‍🏫 Horario Profesores")
    ws5.sheet_view.showGridLines = False

    ws5.merge_cells("A1:H1")
    c = ws5["A1"]
    c.value     = "Horario Semanal por Profesor — Clases Asignadas"
    c.font      = Font(name="Arial", bold=True, size=13, color="FFFFFF")
    c.fill      = PatternFill("solid", fgColor=C["azul_dark"])
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws5.row_dimensions[1].height = 26

    for ci, h in enumerate(["Profesor","Tipo","Día","Horario","Materia","Grupo","Salón","Bloque"], 1):
        _hc(ws5, 2, ci, h, bg=C["azul_med"])
    ws5.row_dimensions[2].height = 18

    prof_clases: dict = defaultdict(list)
    for a in asignadas:
        prof_clases[a["profesor"]].append(a)

    cur_row = 3
    for prof in sorted(prof_clases.keys()):
        cls  = sorted(prof_clases[prof], key=lambda x: (
            DIAS_SEMANA.index(x["dia_asignado"]),
            parsear_horario(x["horario"])[0]
        ))
        tipo = cls[0]["tipo"]
        for ci in range(1, 9):
            ws5.cell(row=cur_row, column=ci).fill   = PatternFill("solid", fgColor=C["azul_light"])
            ws5.cell(row=cur_row, column=ci).border = _borde()
        ws5.cell(row=cur_row, column=1).value     = prof
        ws5.cell(row=cur_row, column=1).font      = Font(name="Arial", bold=True, size=9, color=C["azul_dark"])
        ws5.cell(row=cur_row, column=1).alignment = Alignment(horizontal="left", vertical="center")
        ws5.cell(row=cur_row, column=2).value     = tipo
        ws5.cell(row=cur_row, column=2).font      = Font(name="Arial", bold=True, size=9, color=C["azul_dark"])
        ws5.cell(row=cur_row, column=2).alignment = Alignment(horizontal="center", vertical="center")
        ws5.cell(row=cur_row, column=3).value     = f"{len(cls)} clase(s)"
        ws5.cell(row=cur_row, column=3).font      = Font(name="Arial", italic=True, size=9, color=C["azul_dark"])
        ws5.row_dimensions[cur_row].height = 15
        cur_row += 1
        for cl in cls:
            dia = cl["dia_asignado"]
            bg  = DIA_COLOR.get(dia, C["blanco"])
            for ci, v in enumerate(["","",dia,cl["horario"],cl["materia"],cl["grupo"],cl["salon_asignado"],cl["bloque_salon"]], 1):
                al = "center" if ci in [1,2,3,4,7] else "left"
                _dc(ws5, cur_row, ci, v, bg=bg, align=al, sz=9)
            ws5.row_dimensions[cur_row].height = 14
            cur_row += 1

    for ci, w in enumerate([24,12,12,14,32,10,10,16], 1):
        ws5.column_dimensions[get_column_letter(ci)].width = w
    ws5.freeze_panes = "A3"

    os.makedirs(os.path.dirname(ruta_salida) or ".", exist_ok=True)
    wb.save(ruta_salida)
    print(f"\n✅ Excel exportado: {ruta_salida}")
    print(f"   Hojas: {[s.title for s in wb.worksheets]}")


def exportar_dashboard(resultado: dict, carpeta: str = "dashboard"):
    """
    Exporta los resultados como data.js dentro de la carpeta 'dashboard',
    lista para ser leída por index.html.
    """
    asig   = resultado["asignadas"]
    noasig = resultado["no_asignadas"]
    total  = len(asig) + len(noasig)
    desps  = [a["desperdicio"] for a in asig]

    por_bloque  = dict(Counter(a["bloque_salon"] for a in asig).most_common())
    por_horario = dict(Counter(a["horario"] for a in asig).most_common(12))

    def clean(obj):
        if isinstance(obj, dict):  return {k: clean(v) for k, v in obj.items()}
        if isinstance(obj, list):  return [clean(v) for v in obj]
        if hasattr(obj, "item"):   return obj.item()
        return obj

    data = {
        "resumen": {
            "total":                total,
            "asignadas":            len(asig),
            "no_asignadas":         len(noasig),
            "pct_exito":            round(len(asig) / total * 100, 1) if total else 0,
            "desperdicio_promedio": round(sum(desps) / len(desps), 1) if desps else 0,
            "desperdicio_min":      int(min(desps)) if desps else 0,
            "desperdicio_max":      int(max(desps)) if desps else 0,
        },
        "por_bloque":   por_bloque,
        "por_horario":  por_horario,
        "asignadas":    clean(asig),
        "no_asignadas": clean(noasig),
    }

    os.makedirs(carpeta, exist_ok=True)
    ruta = os.path.join(carpeta, "data.js")
    with open(ruta, "w", encoding="utf-8") as f:
        f.write("const SMARTCAMPUS_DATA = ")
        json.dump(data, f, ensure_ascii=False, indent=2, default=str)
        f.write(";")

    print(f"✅ Dashboard exportado → {ruta}")
    print(f"   Abre dashboard/index.html en tu navegador")
